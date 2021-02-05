import openpyxl
from openpyxl import load_workbook

def handle_excel(filepath=None):
    if not filepath:
        filepath='sendto.xlsx'
    wb=load_workbook(filepath)
    sheet=wb.active
    i=sheet.max_row

    values=[]

    for ii in range(1,i):
        value=[]
        nick_name=sheet.cell(ii,3).value
        qq=sheet.cell(ii,5).value
        qqemail=str(qq)+'@qq.com'
        value.append(nick_name)
        value.append(qqemail)
        #print(value)

        values.append(value)
    #print(values)
    return values

if __name__=="__main__":
    values=handle_excel()
    print(values)