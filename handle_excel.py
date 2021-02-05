import openpyxl
from openpyxl import load_workbook

def handle_excel(filepath=None):
    if not filepath:
        filepath='语料.xlsx'
    wb=load_workbook(filepath)
    sheet=wb.active
    i=sheet.max_row
    j=sheet.max_column
    yuliao_column=0
    result_column=0

    for jj in range(1,j):
        if sheet.cell(1,jj).value=="语料":
            yuliao_column=jj
        elif sheet.cell(1,jj).value=="结果":
            result_column=jj
    yuliao_values=[]
    for ii in range(2,i+1):
        yuliao_value=sheet.cell(ii,yuliao_column).value
        yuliao_values.append(yuliao_value)
    return yuliao_values

if __name__=="__main__":
    yuliao_values=handle_excel()
    print(yuliao_values)
    print(len(yuliao_values))